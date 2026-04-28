"""
Test the v5 template end-to-end:
  1. Copy the template
  2. Inject 9 test scenarios into rows 6-14 covering valid/invalid combinations
  3. Run LibreOffice headless to recalculate all formulas
  4. Read back the calculated W (Row Status) + X (Missing Fields) per row
  5. Read back the Submission Readiness dashboard counts
  6. Compare actual vs expected, report PASS/FAIL per scenario
"""

import shutil, subprocess, os
from openpyxl import load_workbook
from datetime import datetime

SRC = "/sessions/sweet-clever-archimedes/mnt/outputs/idi_migration/4593_ID_Inquiries_Migration_Intake_v5.xlsx"
TEST_DIR = "/sessions/sweet-clever-archimedes/mnt/outputs/idi_migration/test_v5"
TEST_FILE = f"{TEST_DIR}/v5_test_input.xlsx"
OUT_DIR = f"{TEST_DIR}/recalculated"

os.makedirs(TEST_DIR, exist_ok=True)
os.makedirs(OUT_DIR, exist_ok=True)

# Copy template to a working file
shutil.copy(SRC, TEST_FILE)

wb = load_workbook(TEST_FILE)
ws = wb["Instructions"]

# ----------------------------------------------------------------------
# TEST SCENARIOS — each row tests a specific case
# ----------------------------------------------------------------------
# row, label, expected_W, expected_X_contains, fields
scenarios = [
    {
        "row": 6, "label": "Fully valid On Hold row", "expected_W": "Ready", "expected_X_contains": [],
        "data": {
            "A": "The Swan Inn", "B": "14 Greengate Street", "C": "Stafford",
            "D": "Staffordshire", "E": "ST16 3RW",
            "F": "Staffordshire County Council", "G": "Trust Inns Ltd",
            "H": "Edward Paul Bolton", "I": "On Hold", "J": "Hamlins LLP",
            "K": "Michael Green (Hamlins LLP)", "L": "591/PPLPRS1734",
            "M": datetime(2024, 5, 2), "N": "Pub/Bar – background", "O": "No",
            "P": "Confirm Licensee and DPS", "Q": None, "R": None,
            "S": "Public Register Online", "T": datetime(2024, 5, 7),
            "U": "Yes", "V": "Email about hold",
        }
    },
    {
        "row": 7, "label": "Fully valid Allocated row (with Q + R)", "expected_W": "Ready", "expected_X_contains": [],
        "data": {
            "A": "Station Bar", "B": "12 Camden Road", "C": "London",
            "D": "London", "E": "NW1 8AL", "F": "London Borough of Camden",
            "G": "Boom Music Group Ltd", "H": "Liam O'Hare",
            "I": "Allocated", "J": "Hamlins LLP",
            "K": "Michael Green (Hamlins LLP)", "L": "591/PPLPRS3268",
            "M": datetime(2026, 3, 6), "N": "Pub/Bar – background", "O": "No",
            "P": None, "Q": "Stuart Paton", "R": datetime(2026, 3, 16),
            "S": "Email", "T": datetime(2026, 3, 9), "U": "Awaiting Response", "V": None,
        }
    },
    {
        "row": 8, "label": "Allocated row missing investigator (FAIL)", "expected_W": "Incomplete",
        "expected_X_contains": ["Investigator"],
        "data": {
            "A": "The Red Lion", "B": "5 High Street", "C": "Manchester",
            "D": "Greater Manchester", "E": "M1 1AA",
            "F": "Manchester City Council", "I": "Allocated",
            "J": "Brodies LLP", "K": "Lucy Duff (Brodies LLP)",
            "M": datetime(2026, 4, 1), "N": "Pub/Bar – event", "O": "No",
            "Q": None, "R": datetime(2026, 4, 5),
        }
    },
    {
        "row": 9, "label": "Missing Postcode (FAIL)", "expected_W": "Incomplete",
        "expected_X_contains": ["Postcode"],
        "data": {
            "A": "Bonnie Prince Charlie", "B": "20 Norwich Way", "C": "Norwich",
            "D": "Norfolk", "E": None,  # Postcode missing
            "F": "Norfolk County Council", "I": "Unallocated",
            "J": "Hamlins LLP", "K": "Michael Green (Hamlins LLP)",
            "M": datetime(2024, 5, 28), "N": "Pub/Bar – background", "O": "No",
        }
    },
    {
        "row": 10, "label": "Missing Licensing Authority (FAIL)", "expected_W": "Incomplete",
        "expected_X_contains": ["Licensing Authority"],
        "data": {
            "A": "Harry's Bar", "B": "8 Bridge Street", "C": "Menai Bridge",
            "D": "Anglesey", "E": "LL59 5BD",
            "F": None,  # Licensing Authority missing
            "I": "On Hold", "J": "Hamlins LLP",
            "K": "Kirsty Lawrence (Hamlins LLP)", "M": datetime(2024, 7, 2),
            "N": "Pub/Bar – background", "O": "No",
        }
    },
    {
        "row": 11, "label": "Council Yes but no notes (FAIL)", "expected_W": "Incomplete",
        "expected_X_contains": ["Licence Check Notes"],
        "data": {
            "A": "The Hideout", "B": "1 Queens Road", "C": "Edinburgh",
            "D": "Midlothian", "E": "EH1 2AA",
            "F": "City of Edinburgh Council", "I": "Ready for Allocation",
            "J": "Burness Paull LLP",
            "K": "Other – advise A2Z Cloud", "M": datetime(2026, 1, 15),
            "N": "Cafe", "O": "No", "S": "Email",
            "T": datetime(2026, 1, 20), "U": "Yes", "V": None,  # Notes missing
        }
    },
    {
        "row": 12, "label": "Other firm pick (Ready but flagged)", "expected_W": "Ready",
        "expected_X_contains": [],
        "data": {
            "A": "The Crown Pub", "B": "33 Market Street", "C": "Birmingham",
            "D": "West Midlands", "E": "B1 1AA",
            "F": "Birmingham City Council", "I": "Unallocated",
            "J": "Other – advise A2Z Cloud",  # Other firm
            "K": "Other – advise A2Z Cloud",  # Other client
            "M": datetime(2026, 4, 10), "N": "Pub/Bar – event", "O": "Yes",
        }
    },
    {
        "row": 13, "label": "Empty row", "expected_W": "", "expected_X_contains": [],
        "data": {}  # all blank
    },
    {
        "row": 14, "label": "Multiple missing required fields (FAIL)",
        "expected_W": "Incomplete",
        "expected_X_contains": ["Street Address", "Postcode", "Licensing Authority"],
        "data": {
            "A": "Some Premises",
            "I": "Unallocated", "J": "Hamlins LLP",
            "K": "Michael Green (Hamlins LLP)",
            "M": datetime(2026, 4, 20), "N": "Cafe", "O": "No",
        }
    },
]

print("=" * 75)
print("INJECTING TEST DATA")
print("=" * 75)

for sc in scenarios:
    r = sc["row"]
    for col, val in sc["data"].items():
        ws[f"{col}{r}"] = val
    print(f"  Row {r}: {sc['label']}")

wb.save(TEST_FILE)
print(f"\nSaved test file: {TEST_FILE}")

# ----------------------------------------------------------------------
# Recalculate via LibreOffice headless
# ----------------------------------------------------------------------
print("\n" + "=" * 75)
print("RECALCULATING via LibreOffice headless")
print("=" * 75)

# Use a unique user profile to avoid collisions
profile = f"{TEST_DIR}/lo_profile"
os.makedirs(profile, exist_ok=True)

cmd = [
    "libreoffice",
    f"-env:UserInstallation=file://{profile}",
    "--headless",
    "--calc",
    "--convert-to", "xlsx",
    "--outdir", OUT_DIR,
    TEST_FILE,
]
result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
print("stdout:", result.stdout.strip())
print("stderr:", result.stderr.strip()[:300])

# Output filename matches input basename
recalc_file = f"{OUT_DIR}/{os.path.basename(TEST_FILE)}"
if not os.path.exists(recalc_file):
    print(f"\nFAIL: recalc file not produced at {recalc_file}")
    print(f"  Files in {OUT_DIR}: {os.listdir(OUT_DIR)}")
    raise SystemExit(1)

print(f"Recalc file: {recalc_file}")

# ----------------------------------------------------------------------
# Read back with cached values
# ----------------------------------------------------------------------
wb2 = load_workbook(recalc_file, data_only=True)
ws2 = wb2["Instructions"]
ready = wb2["Submission Readiness"]

print("\n" + "=" * 75)
print("RESULTS — per-row W (Row Status) and X (Missing Fields)")
print("=" * 75)

passes, fails = 0, 0
for sc in scenarios:
    r = sc["row"]
    actual_w = ws2[f"W{r}"].value
    actual_x = ws2[f"X{r}"].value or ""
    expected_w = sc["expected_W"]
    expected_x_contains = sc["expected_X_contains"]

    # Normalize None to ""
    actual_w_norm = "" if actual_w is None else actual_w

    w_ok = actual_w_norm == expected_w
    x_ok = all(needle in actual_x for needle in expected_x_contains)
    overall = "PASS" if (w_ok and x_ok) else "FAIL"
    if overall == "PASS":
        passes += 1
    else:
        fails += 1

    print(f"\n  Row {r}: {sc['label']}")
    print(f"    Expected W: {expected_w!r:30s}  Actual W: {actual_w!r}")
    if expected_x_contains:
        print(f"    Expected X to contain: {expected_x_contains}")
    print(f"    Actual X: {actual_x!r}")
    print(f"    => {overall}  (W match: {w_ok}, X match: {x_ok})")

print("\n" + "=" * 75)
print("SUBMISSION READINESS DASHBOARD")
print("=" * 75)

# Banner
banner = ready["B2"].value
print(f"\n  Banner (B2): {banner!r}")

# Headline counts
total_rows = ready["C5"].value
valid_rows = ready["C6"].value
incomplete = ready["C7"].value
print(f"\n  Total rows entered:  {total_rows}")
print(f"  Rows fully valid:    {valid_rows}")
print(f"  Rows incomplete:     {incomplete}")

# Issue breakdown
print("\n  Issue breakdown:")
for r in range(11, 21):
    label = ready[f"B{r}"].value
    count = ready[f"C{r}"].value
    example = ready[f"D{r}"].value
    if label:
        print(f"    {label:50s}  count={count:>3}  example_row={example}")

# ----------------------------------------------------------------------
# Expected dashboard values
# ----------------------------------------------------------------------
expected_total = 8  # 9 scenarios but row 13 is empty
expected_valid = 3  # rows 6, 7, 12
expected_incomplete = 5  # rows 8, 9, 10, 11, 14

print("\n" + "=" * 75)
print("DASHBOARD VERIFICATION")
print("=" * 75)
checks = [
    ("Total rows", total_rows, expected_total),
    ("Valid rows", valid_rows, expected_valid),
    ("Incomplete rows", incomplete, expected_incomplete),
]
for name, actual, expected in checks:
    status = "PASS" if actual == expected else "FAIL"
    print(f"  {name:18s}  expected={expected}  actual={actual}  => {status}")
    if status == "PASS":
        passes += 1
    else:
        fails += 1

# Banner check — should be NOT READY since incomplete > 0
expected_banner_substr = "NOT READY"
banner_ok = expected_banner_substr in (banner or "")
status = "PASS" if banner_ok else "FAIL"
print(f"  Banner shows NOT READY (since incomplete>0)  => {status}")
print(f"    actual: {banner!r}")
if banner_ok:
    passes += 1
else:
    fails += 1

print("\n" + "=" * 75)
print(f"OVERALL:  {passes} PASS,  {fails} FAIL")
print("=" * 75)
