"""
Build the ID Inquiries Migration Intake Template v5.

Outputs: 4593_ID_Inquiries_Migration_Intake_v5.xlsx

Sheets:
  1. Cover               - what this file is + 3-step instructions
  2. Guide               - detailed how-to-fill guidance
  3. Instructions        - main data entry (22 cols + 2 diagnostic cols, 400 rows)
  4. Portal Users        - separate intake for portal user emails
  5. Submission Readiness - live dashboard with READY TO SEND indicator
  6. Picklists           - reference data (hidden, locked)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.utils import get_column_letter

# ============================================================
#   REFERENCE DATA — sourced from picklist_report.json + v4
# ============================================================

FIRMS = [
    "Hamlins LLP", "Seddons-GSC", "Brodies LLP", "Burness Paull LLP",
    "Other – advise A2Z Cloud",
]

CLIENTS = [
    "Michael Green (Hamlins LLP)",
    "Kirsty Lawrence (Hamlins LLP)",
    "Wendy Farrow (Hamlins LLP)",
    "Martin Ochs (Hamlins LLP)",
    "Renee Afambu (Hamlins LLP)",
    "Lucy Duff (Brodies LLP)",
    "Other – advise A2Z Cloud",
]

INVESTIGATORS = [
    "Mathew MacMillan", "Dan Blowes", "David Garside", "Gary Hall",
    "Ian Hodgson", "Jane Dowson", "Justine Ranson", "Ken Farrell",
    "Mandi Astill", "Neil Beattie", "Stuart Paton", "Tom Kell",
    "Charles Clawson", "Luke Terry", "Adam McLean", "John McGowan",
    "Other – advise A2Z Cloud",
]

INSTRUCTION_STATUS = [
    "On Hold", "Unallocated", "Ready for Allocation", "Allocated",
]

INSTRUCTION_TYPE = [
    "Pub/Bar – background", "Pub/Bar – event", "Restaurant//Café", "Cafe",
    "Nightclub", "Stripclub", "Gentleman's Club", "Shop",
    "Shop – High Street", "Shop – Warehouse/Industrial", "Beauty Salon",
    "Barber (Gents)", "Hairdresser", "Nail Salon", "Tattoo Parlour", "Gym",
    "Hotel – Bedroom", "Hotel – Common Areas", "Festival", "Other",
]

YES_NO = ["Yes", "No"]
COUNCIL_METHOD = ["Public Register Online", "Email", "Form Submitted"]
COUNCIL_RESPONSE = ["Yes", "No", "Awaiting Response"]
ACTIVATE_PORTAL = ["Yes", "Defer"]

# ============================================================
#   STYLES
# ============================================================

NAVY = "1F3864"
LIGHT_NAVY = "2E5395"
ACCENT = "C00000"      # red for required / errors
SOFT_RED = "FCE4E4"
SOFT_AMBER = "FFF2CC"
SOFT_GREEN = "E2EFDA"
HEADER_GREY = "F2F2F2"
EXAMPLE_BLUE = "DDEBF7"
COVER_BG = "FFFFFF"

WHITE = "FFFFFF"

font_title = Font(name="Calibri", size=22, bold=True, color=NAVY)
font_subtitle = Font(name="Calibri", size=12, italic=True, color=LIGHT_NAVY)
font_section = Font(name="Calibri", size=14, bold=True, color=NAVY)
font_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_subheader = Font(name="Calibri", size=10, bold=True, color=NAVY)
font_body = Font(name="Calibri", size=11)
font_body_bold = Font(name="Calibri", size=11, bold=True)
font_required = Font(name="Calibri", size=11, bold=True, color=ACCENT)
font_descr = Font(name="Calibri", size=9, italic=True, color="595959")
font_example = Font(name="Calibri", size=10, color="305496")
font_diag = Font(name="Calibri", size=10, italic=True, color="595959")
font_big = Font(name="Calibri", size=20, bold=True, color=WHITE)

fill_header = PatternFill("solid", fgColor=NAVY)
fill_subheader = PatternFill("solid", fgColor=HEADER_GREY)
fill_example = PatternFill("solid", fgColor=EXAMPLE_BLUE)
fill_amber = PatternFill("solid", fgColor=SOFT_AMBER)
fill_green = PatternFill("solid", fgColor=SOFT_GREEN)
fill_red = PatternFill("solid", fgColor=SOFT_RED)
fill_navy = PatternFill("solid", fgColor=NAVY)
fill_light_navy = PatternFill("solid", fgColor=LIGHT_NAVY)
fill_white = PatternFill("solid", fgColor=WHITE)

thin = Side(border_style="thin", color="BFBFBF")
thick = Side(border_style="medium", color=NAVY)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
border_header = Border(left=thin, right=thin, top=thick, bottom=thick)

align_centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
align_centre_top = Alignment(horizontal="center", vertical="top", wrap_text=True)


# ============================================================
#   COLUMN SCHEMA — Instructions sheet
# ============================================================
# tuples: (col_letter, section, header, required, descr, picklist, width)

COLUMNS = [
    # PREMISES (A-H)
    ("A", "PREMISES",     "Premises Name",            True,  "Name of the licensed venue exactly as on the licence.", None, 28),
    ("B", "PREMISES",     "Street Address",           True,  "Full street address (was missing in test extract — please backfill).", None, 30),
    ("C", "PREMISES",     "City",                     True,  "Town or city.", None, 16),
    ("D", "PREMISES",     "County",                   False, "County (optional).", None, 16),
    ("E", "PREMISES",     "Postcode",                 True,  "Full UK postcode.", None, 11),
    ("F", "PREMISES",     "Licensing Authority",      True,  "Council that issued the licence (was missing in test extract — please backfill).", None, 26),
    ("G", "PREMISES",     "Premises Licence Holder",  False, "Legal entity holding the premises licence.", None, 24),
    ("H", "PREMISES",     "DPS",                      False, "Designated Premises Supervisor full name.", None, 22),
    # INSTRUCTION (I-P)
    ("I", "INSTRUCTION",  "Instruction Status",       True,  "▼ Current status of this instruction.", "InstructionStatus", 18),
    ("J", "INSTRUCTION",  "Instructing Company",      True,  "▼ Law firm. Pick 'Other' if not listed.", "Firms", 22),
    ("K", "INSTRUCTION",  "Instructing Client",       True,  "▼ Individual lawyer (firm shown in brackets).", "Clients", 28),
    ("L", "INSTRUCTION",  "Client Ref",               False, "Client's own reference number.", None, 18),
    ("M", "INSTRUCTION",  "Date of Instruction",      True,  "Date instruction received. DD/MM/YYYY.", None, 14),
    ("N", "INSTRUCTION",  "Instruction Type",         True,  "▼ Venue type closest to this premises. (Replaces 'Basic'.)", "InstructionType", 22),
    ("O", "INSTRUCTION",  "Is Reinstruction?",        True,  "▼ Repeat visit to same premises?", "YesNo", 12),
    ("P", "INSTRUCTION",  "Instruction Notes",        False, "Special instructions for the investigator — copy from your source notes.", None, 36),
    # INVESTIGATOR (Q-R)
    ("Q", "INVESTIGATOR", "Investigator Assigned",    False, "▼ Required IF Status = Allocated. Leave blank otherwise.", "Investigators", 22),
    ("R", "INVESTIGATOR", "Date Assigned",            False, "Required IF Status = Allocated. DD/MM/YYYY.", None, 14),
    # LICENCE CHECK (S-V)
    ("S", "LICENCE CHECK","Council Request Method",   False, "▼ How licence details were requested.", "CouncilMethod", 22),
    ("T", "LICENCE CHECK","Council Request Date",     False, "Date council was contacted. DD/MM/YYYY.", None, 14),
    ("U", "LICENCE CHECK","Council Response?",        False, "▼ Has the council responded?", "CouncilResponse", 18),
    ("V", "LICENCE CHECK","Licence Check Notes",      False, "Required IF Council Response = Yes.", None, 36),
    # DIAGNOSTICS (W-X) — formula-driven
    ("W", "DIAGNOSTICS",  "Row Status",               False, "Auto-calculated.", None, 14),
    ("X", "DIAGNOSTICS",  "Missing Fields",           False, "Auto-calculated. Lists what's missing.", None, 40),
]

SECTION_COLOURS = {
    "PREMISES":     "1F3864",
    "INSTRUCTION":  "2E5395",
    "INVESTIGATOR": "548235",
    "LICENCE CHECK":"BF8F00",
    "DIAGNOSTICS":  "595959",
}

DATA_FIRST_ROW = 6   # row 1=section, row 2=header, row 3=descr, row 4=example, row 5=blank divider, row 6=first real
DATA_LAST_ROW = 405  # 400 rows
EXAMPLE_ROW = 4

EXAMPLE_VALUES = {
    "A": "The Swan Inn",
    "B": "14 Greengate Street",
    "C": "Stafford",
    "D": "Staffordshire",
    "E": "ST16 3RW",
    "F": "Staffordshire County Council",
    "G": "Trust Inns Ltd",
    "H": "Edward Paul Bolton",
    "I": "On Hold",
    "J": "Hamlins LLP",
    "K": "Michael Green (Hamlins LLP)",
    "L": "591/PPLPRS1734",
    "M": "02/05/2024",
    "N": "Pub/Bar – background",
    "O": "No",
    "P": "Confirm Licensee and DPS. Attempt to obtain bank details.",
    "Q": "Gary Hall",
    "R": "16/05/2024",
    "S": "Public Register Online",
    "T": "07/05/2024",
    "U": "Yes",
    "V": "Email from Michael asking to put on hold — linked premises found.",
}


# ============================================================
#   BUILD
# ============================================================

def unlock(cell):
    cell.protection = Protection(locked=False)


def build_picklists(wb):
    ws = wb.create_sheet("Picklists")
    ws.sheet_state = "hidden"

    columns = [
        ("Firms",               FIRMS),
        ("Clients",             CLIENTS),
        ("Investigators",       INVESTIGATORS),
        ("InstructionStatus",   INSTRUCTION_STATUS),
        ("InstructionType",     INSTRUCTION_TYPE),
        ("YesNo",               YES_NO),
        ("CouncilMethod",       COUNCIL_METHOD),
        ("CouncilResponse",     COUNCIL_RESPONSE),
        ("ActivatePortal",      ACTIVATE_PORTAL),
    ]

    for idx, (name, values) in enumerate(columns, start=1):
        col = get_column_letter(idx)
        ws[f"{col}1"] = name
        ws[f"{col}1"].font = font_subheader
        ws[f"{col}1"].fill = fill_subheader
        for row, v in enumerate(values, start=2):
            ws[f"{col}{row}"] = v
        ws.column_dimensions[col].width = max(len(name), max((len(v) for v in values), default=10)) + 2

        # Defined name for each picklist for use in DataValidation
        from openpyxl.workbook.defined_name import DefinedName
        last_row = 1 + len(values)
        ref = f"Picklists!${col}$2:${col}${last_row}"
        defn = DefinedName(name=name, attr_text=ref)
        wb.defined_names[name] = defn

    ws.protection.sheet = True
    return ws


def build_cover(wb):
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 110
    ws.column_dimensions["C"].width = 4

    ws["B2"] = "ID Inquiries Migration — Data Collection Template"
    ws["B2"].font = font_title
    ws["B2"].alignment = align_left

    ws["B3"] = "Project #4593   |   Version 5.1   |   April 2026   |   Prepared by A2Z Cloud"
    ws["B3"].font = font_subtitle
    ws["B3"].alignment = align_left

    ws.row_dimensions[2].height = 36

    ws["B5"] = "Welcome"
    ws["B5"].font = font_section

    ws["B6"] = (
        "This single workbook captures everything we need to migrate ID Inquiries' active instructions into the new "
        "Zoho CRM. You only need to fill in two tabs — Instructions and Portal Users — and check the Submission "
        "Readiness tab tells you when the file is ready to send back."
    )
    ws["B6"].font = font_body
    ws["B6"].alignment = align_left_top
    ws.row_dimensions[6].height = 50

    ws["B8"] = "How to use this file"
    ws["B8"].font = font_section

    steps = [
        ("Step 1", "Read the Guide tab — it explains every column and what's required."),
        ("Step 2", "Go to the Instructions tab and fill one row per active instruction. Required cells turn red if left blank."),
        ("Step 3", "Fill the Portal Users tab if you have client portal users to set up. (Optional — you can do this later.)"),
        ("Step 4", "Open the Submission Readiness tab. When it shows READY TO SEND: YES, email this file back to A2Z Cloud."),
    ]
    for i, (label, text) in enumerate(steps):
        r = 10 + i * 2
        ws[f"B{r}"] = f"{label} — {text}"
        ws[f"B{r}"].font = font_body
        ws[f"B{r}"].alignment = align_left_top
        ws.row_dimensions[r].height = 30

    ws["B19"] = "Status legend"
    ws["B19"].font = font_section

    legend = [
        ("✓ Ready",       "Row is fully valid and will import cleanly.",                fill_green),
        ("! Incomplete",  "Required cells are empty — see the Missing Fields column.",  fill_amber),
        ("× Other pick",  "You picked 'Other – advise A2Z Cloud' — flag for our team.", fill_red),
    ]
    for i, (label, text, fill) in enumerate(legend):
        r = 21 + i
        ws[f"B{r}"] = f"  {label}    {text}"
        ws[f"B{r}"].font = font_body
        ws[f"B{r}"].fill = fill
        ws[f"B{r}"].alignment = align_left
        ws.row_dimensions[r].height = 22

    ws["B26"] = "Contact"
    ws["B26"].font = font_section
    ws["B27"] = "Sam — A2Z Cloud, Zoho Development Team"
    ws["B27"].font = font_body
    ws["B28"] = "Project #4593 — ID Inquiries Ltd"
    ws["B28"].font = font_body

    ws.protection.sheet = True
    return ws


def build_guide(wb):
    ws = wb.create_sheet("Guide")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 80

    ws["B2"] = "How to fill this template"
    ws["B2"].font = font_title

    sections = [
        ("Required vs optional",
         "Columns marked with * in the header (and shown in red) are required for every row. Columns without * are optional."),
        ("Conditionally required",
         "If Instruction Status = Allocated, then Investigator Assigned + Date Assigned are required for that row.\n"
         "If Council Response = Yes, then Licence Check Notes are required for that row."),
        ("Dropdowns (▼ icon)",
         "Click the cell, then click the dropdown arrow that appears, and pick a value. Don't type values — typed values "
         "will be flagged as invalid in the Submission Readiness tab."),
        ("'Other – advise A2Z Cloud'",
         "If a firm, client, or investigator isn't in the dropdown, pick this option and write a comment in the next "
         "column. We'll add the missing record before importing."),
        ("Dates",
         "Type dates as DD/MM/YYYY (e.g. 16/05/2024). The cell format is locked to date — Excel may reformat your input."),
        ("'Basic' instruction type",
         "The old extract used 'Basic' as a catch-all. The new template requires you to pick the actual venue type "
         "(Pub/Bar – background, Restaurant//Café, etc.) from the dropdown — this matches what the CRM expects."),
        ("Blank cells",
         "Leave optional cells genuinely empty. Don't type 'N/A', 'None', or '-'. The validation treats those as valid "
         "values and they'll end up in the CRM."),
        ("Submission Readiness",
         "Open this tab any time to see how many rows are ready, what's missing, and whether you can send the file back. "
         "It updates automatically as you fill rows."),
        ("When you're done",
         "Submission Readiness shows READY TO SEND: YES → email the file back to Sam at A2Z Cloud."),
    ]
    for i, (h, body) in enumerate(sections):
        r = 4 + i * 2
        ws[f"B{r}"] = h
        ws[f"B{r}"].font = font_body_bold
        ws[f"C{r}"] = body
        ws[f"C{r}"].font = font_body
        ws[f"C{r}"].alignment = align_left_top
        ws.row_dimensions[r].height = 36 if "\n" in body else 26

    ws.protection.sheet = True
    return ws


def build_instructions(wb):
    ws = wb.create_sheet("Instructions")
    ws.sheet_view.showGridLines = False

    # Row 1: section group bands
    sections_grouped = []
    last_section = None
    start = None
    for col_letter, section, *_ in COLUMNS:
        if section != last_section:
            if last_section is not None:
                sections_grouped.append((last_section, start, prev_col))
            last_section = section
            start = col_letter
        prev_col = col_letter
    sections_grouped.append((last_section, start, prev_col))

    for section, c0, c1 in sections_grouped:
        ws.merge_cells(f"{c0}1:{c1}1")
        cell = ws[f"{c0}1"]
        cell.value = section
        cell.font = font_header
        cell.fill = PatternFill("solid", fgColor=SECTION_COLOURS[section])
        cell.alignment = align_centre
        cell.border = border_header
    ws.row_dimensions[1].height = 24

    # Row 2: column headers (with * if required)
    for col_letter, _, header, required, _, _, width in COLUMNS:
        cell = ws[f"{col_letter}2"]
        cell.value = f"{header} *" if required else header
        cell.font = font_required if required else font_subheader
        cell.fill = fill_subheader
        cell.alignment = align_centre
        cell.border = border_all
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[2].height = 32

    # Row 3: descriptions
    for col_letter, _, _, _, descr, _, _ in COLUMNS:
        cell = ws[f"{col_letter}3"]
        cell.value = descr
        cell.font = font_descr
        cell.alignment = align_left_top
        cell.border = border_all
    ws.row_dimensions[3].height = 50

    # Row 4: example row
    for col_letter, _, _, _, _, _, _ in COLUMNS:
        cell = ws[f"{col_letter}{EXAMPLE_ROW}"]
        cell.value = EXAMPLE_VALUES.get(col_letter, "")
        cell.font = font_example
        cell.fill = fill_example
        cell.alignment = align_left
        cell.border = border_all
    # Example row label
    ws[f"A{EXAMPLE_ROW}"].comment = None
    ws.row_dimensions[EXAMPLE_ROW].height = 20

    # Row 5: hint row "▶ EXAMPLE ABOVE — START YOUR DATA FROM ROW 6"
    ws.merge_cells(f"A5:V5")
    hint = ws["A5"]
    hint.value = "▶  Row 4 is an EXAMPLE — leave it as a reference. Start filling your real data from row 6 below."
    hint.font = font_descr
    hint.fill = fill_amber
    hint.alignment = align_centre
    ws.row_dimensions[5].height = 22

    # Diagnostic columns formulas in W and X for rows 6..405
    # W: Row Status — "" if empty, "Incomplete" or "Ready"
    # X: Missing Fields — TEXTJOIN of missing required field names
    REQUIRED_COLS = [c for c, _, _, req, _, _, _ in COLUMNS if req]

    for r in range(DATA_FIRST_ROW, DATA_LAST_ROW + 1):
        # Row status formula
        empty_check = '=IF(A{r}="","",'
        # required-field check
        cond_parts = [f'A{r}=""']
        for c in ["B", "C", "E", "F", "I", "J", "K", "M", "N", "O"]:
            cond_parts.append(f'{c}{r}=""')
        # conditional: if I="Allocated" then Q and R must be non-empty
        cond_parts.append(f'AND(I{r}="Allocated",OR(Q{r}="",R{r}=""))')
        # conditional: if U="Yes" then V must be non-empty
        cond_parts.append(f'AND(U{r}="Yes",V{r}="")')
        condition = "OR(" + ",".join(cond_parts) + ')'
        formula_w = f'=IF(A{r}="","",IF({condition},"Incomplete","Ready"))'
        ws[f"W{r}"] = formula_w
        ws[f"W{r}"].font = font_diag
        ws[f"W{r}"].alignment = align_centre

        # Missing fields formula — only populated when row is Incomplete.
        # Uses & concat + SUBSTITUTE to strip the leading ", " separator.
        # Compatible with Excel 2007+, Excel for Mac, LibreOffice, Google Sheets.
        # Each IF returns ", <Label>" or "" — SUBSTITUTE removes the leading separator.
        miss_parts = []
        labels = {
            "B": "Street Address", "C": "City", "E": "Postcode",
            "F": "Licensing Authority", "I": "Status", "J": "Firm",
            "K": "Client", "M": "Date of Instruction", "N": "Instruction Type",
            "O": "Reinstruction",
        }
        for col, lbl in labels.items():
            miss_parts.append(f'IF({col}{r}="",", {lbl}","")')
        miss_parts.append(f'IF(AND(I{r}="Allocated",Q{r}=""),", Investigator","")')
        miss_parts.append(f'IF(AND(I{r}="Allocated",R{r}=""),", Date Assigned","")')
        miss_parts.append(f'IF(AND(U{r}="Yes",V{r}=""),", Licence Check Notes","")')
        concat = "&".join(miss_parts)
        formula_x = f'=IF(W{r}<>"Incomplete","",SUBSTITUTE({concat},", ","",1))'
        ws[f"X{r}"] = formula_x
        ws[f"X{r}"].font = font_diag
        ws[f"X{r}"].alignment = align_left

    # Unlock data entry cells (A6:V405)
    for r in range(DATA_FIRST_ROW, DATA_LAST_ROW + 1):
        for col_letter in [c for c, _, _, _, _, _, _ in COLUMNS if c not in ("W", "X")]:
            unlock(ws[f"{col_letter}{r}"])

    # Data validations
    pl_map = {
        "InstructionStatus": "InstructionStatus",
        "Firms": "Firms",
        "Clients": "Clients",
        "InstructionType": "InstructionType",
        "YesNo": "YesNo",
        "Investigators": "Investigators",
        "CouncilMethod": "CouncilMethod",
        "CouncilResponse": "CouncilResponse",
    }

    for col_letter, _, _, _, _, picklist, _ in COLUMNS:
        if picklist is None:
            continue
        defined = pl_map[picklist]
        dv = DataValidation(
            type="list",
            formula1=f"={defined}",
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid value",
            error=f"Pick from the dropdown list. If the value you need isn't there, choose 'Other – advise A2Z Cloud'.",
        )
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}{DATA_FIRST_ROW}:{col_letter}{DATA_LAST_ROW}")

    # Date validations on M, R, T
    for col in ["M", "R", "T"]:
        dv = DataValidation(
            type="date",
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Date format",
            error="Please enter a valid date in DD/MM/YYYY format.",
        )
        ws.add_data_validation(dv)
        dv.add(f"{col}{DATA_FIRST_ROW}:{col}{DATA_LAST_ROW}")

    # Conditional formatting
    # 1. Required cell empty BUT row has been started (Premises Name filled) → red fill
    REQUIRED_COLS_INC_FIRST = ["A", "B", "C", "E", "F", "I", "J", "K", "M", "N", "O"]
    for col in REQUIRED_COLS_INC_FIRST:
        rule = FormulaRule(
            formula=[f'AND(${col}{DATA_FIRST_ROW}="",$A{DATA_FIRST_ROW}<>"",ROW()>={DATA_FIRST_ROW})'],
            stopIfTrue=False,
            fill=fill_red,
        )
        ws.conditional_formatting.add(
            f"{col}{DATA_FIRST_ROW}:{col}{DATA_LAST_ROW}", rule
        )

    # 2. Conditional required: Allocated → Q/R must be filled
    for col in ["Q", "R"]:
        rule = FormulaRule(
            formula=[f'AND(${col}{DATA_FIRST_ROW}="",$I{DATA_FIRST_ROW}="Allocated")'],
            stopIfTrue=False,
            fill=fill_red,
        )
        ws.conditional_formatting.add(f"{col}{DATA_FIRST_ROW}:{col}{DATA_LAST_ROW}", rule)

    # 3. Conditional required: Council Response = Yes → V required
    rule = FormulaRule(
        formula=[f'AND($V{DATA_FIRST_ROW}="",$U{DATA_FIRST_ROW}="Yes")'],
        stopIfTrue=False,
        fill=fill_red,
    )
    ws.conditional_formatting.add(f"V{DATA_FIRST_ROW}:V{DATA_LAST_ROW}", rule)

    # 4. Row status colours — green for Ready, amber for Incomplete
    rule_ready = FormulaRule(
        formula=[f'$W{DATA_FIRST_ROW}="Ready"'], stopIfTrue=False, fill=fill_green,
    )
    rule_inc = FormulaRule(
        formula=[f'$W{DATA_FIRST_ROW}="Incomplete"'], stopIfTrue=False, fill=fill_amber,
    )
    ws.conditional_formatting.add(f"W{DATA_FIRST_ROW}:W{DATA_LAST_ROW}", rule_ready)
    ws.conditional_formatting.add(f"W{DATA_FIRST_ROW}:W{DATA_LAST_ROW}", rule_inc)

    # 5. "Other" picks — amber on the picked cell
    for col in ["J", "K", "Q"]:
        rule_other = FormulaRule(
            formula=[f'${col}{DATA_FIRST_ROW}="Other – advise A2Z Cloud"'],
            stopIfTrue=False, fill=fill_amber,
        )
        ws.conditional_formatting.add(
            f"{col}{DATA_FIRST_ROW}:{col}{DATA_LAST_ROW}", rule_other,
        )

    # Freeze panes — keep first 3 rows + col A visible
    ws.freeze_panes = "B6"

    # Sheet protection
    ws.protection.sheet = True
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertRows = False
    return ws


def build_portal_users(wb):
    ws = wb.create_sheet("Portal Users")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"] = "Portal Users — clients who need login access to submit instructions"
    ws["A1"].font = font_header
    ws["A1"].fill = fill_navy
    ws["A1"].alignment = align_centre
    ws.row_dimensions[1].height = 26

    headers = [
        ("A", "Client Name", 32, "Clients"),
        ("B", "Firm",        24, None),
        ("C", "Email",       30, None),
        ("D", "Phone",       18, None),
        ("E", "Activate Portal?", 16, "ActivatePortal"),
    ]
    for col, h, w, _ in headers:
        cell = ws[f"{col}2"]
        cell.value = f"{h} *" if h in ("Client Name", "Email") else h
        cell.font = font_required if h in ("Client Name", "Email") else font_subheader
        cell.fill = fill_subheader
        cell.alignment = align_centre
        cell.border = border_all
        ws.column_dimensions[col].width = w
    ws.row_dimensions[2].height = 26

    # Description row
    descrs = {
        "A": "▼ Pick from dropdown.",
        "B": "Auto-fill — leave blank.",
        "C": "REQUIRED. Their direct email.",
        "D": "Optional.",
        "E": "▼ Yes = activate now.  Defer = set up later.",
    }
    for col, txt in descrs.items():
        cell = ws[f"{col}3"]
        cell.value = txt
        cell.font = font_descr
        cell.alignment = align_left_top
        cell.border = border_all
    ws.row_dimensions[3].height = 32

    # 50 fillable rows from row 4
    portal_first, portal_last = 4, 53
    for r in range(portal_first, portal_last + 1):
        for col, _, _, _ in headers:
            unlock(ws[f"{col}{r}"])

    dv_clients = DataValidation(type="list", formula1="=Clients", allow_blank=True)
    ws.add_data_validation(dv_clients)
    dv_clients.add(f"A{portal_first}:A{portal_last}")

    dv_act = DataValidation(type="list", formula1="=ActivatePortal", allow_blank=True)
    ws.add_data_validation(dv_act)
    dv_act.add(f"E{portal_first}:E{portal_last}")

    ws.freeze_panes = "A4"
    ws.protection.sheet = True
    return ws


def build_readiness(wb):
    ws = wb.create_sheet("Submission Readiness")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 60

    # Big READY TO SEND banner
    ws.merge_cells("B2:D2")
    ws["B2"] = '=IF(AND(C5>0,C7=0),"READY TO SEND ✓  —  Email this file back to A2Z Cloud","NOT READY  ×  —  Fix the issues below before sending")'
    ws["B2"].font = font_big
    ws["B2"].alignment = align_centre
    ws.row_dimensions[2].height = 50

    # CF on the banner: green if ready, red if not
    rule_green_banner = FormulaRule(
        formula=['AND(C5>0,C7=0)'], stopIfTrue=False, fill=fill_navy,
    )
    rule_red_banner = FormulaRule(
        formula=['NOT(AND(C5>0,C7=0))'], stopIfTrue=False,
        fill=PatternFill("solid", fgColor=ACCENT),
    )
    ws.conditional_formatting.add("B2:D2", rule_green_banner)
    ws.conditional_formatting.add("B2:D2", rule_red_banner)

    # Headline counts
    ws["B4"] = "Headline counts"
    ws["B4"].font = font_section

    counts = [
        ("Total rows entered",
         '=COUNTA(Instructions!A6:A405)'),
        ("Rows fully valid (Ready)",
         '=COUNTIF(Instructions!W6:W405,"Ready")'),
        ("Rows incomplete",
         '=COUNTIF(Instructions!W6:W405,"Incomplete")'),
    ]
    for i, (label, formula) in enumerate(counts):
        r = 5 + i
        ws[f"B{r}"] = label
        ws[f"B{r}"].font = font_body
        ws[f"C{r}"] = formula
        ws[f"C{r}"].font = font_body_bold
        ws[f"C{r}"].alignment = align_centre
        ws[f"B{r}"].border = border_all
        ws[f"C{r}"].border = border_all
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 22
    ws.row_dimensions[7].height = 22

    # Issue breakdown
    ws["B9"] = "Issue breakdown"
    ws["B9"].font = font_section

    headers = ["Issue", "Count", "First example row"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=10, column=2 + i, value=h)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_centre
        cell.border = border_all
    ws.row_dimensions[10].height = 22

    issues = [
        ("Missing Street Address",
         '=SUMPRODUCT((Instructions!A6:A405<>"")*(Instructions!B6:B405=""))',
         '=IFERROR(MATCH(1,(Instructions!A6:A405<>"")*(Instructions!B6:B405=""),0)+5,"—")'),
        ("Missing Licensing Authority",
         '=SUMPRODUCT((Instructions!A6:A405<>"")*(Instructions!F6:F405=""))',
         '=IFERROR(MATCH(1,(Instructions!A6:A405<>"")*(Instructions!F6:F405=""),0)+5,"—")'),
        ("Missing Postcode",
         '=SUMPRODUCT((Instructions!A6:A405<>"")*(Instructions!E6:E405=""))',
         '=IFERROR(MATCH(1,(Instructions!A6:A405<>"")*(Instructions!E6:E405=""),0)+5,"—")'),
        ("Missing Status / Firm / Client",
         '=SUMPRODUCT((Instructions!A6:A405<>"")*((Instructions!I6:I405="")+(Instructions!J6:J405="")+(Instructions!K6:K405="")>0))',
         '=IFERROR(MATCH(1,(Instructions!A6:A405<>"")*((Instructions!I6:I405="")+(Instructions!J6:J405="")+(Instructions!K6:K405="")>0),0)+5,"—")'),
        ("Missing Date / Type / Reinstruction flag",
         '=SUMPRODUCT((Instructions!A6:A405<>"")*((Instructions!M6:M405="")+(Instructions!N6:N405="")+(Instructions!O6:O405="")>0))',
         '=IFERROR(MATCH(1,(Instructions!A6:A405<>"")*((Instructions!M6:M405="")+(Instructions!N6:N405="")+(Instructions!O6:O405="")>0),0)+5,"—")'),
        ("Allocated rows missing investigator/date",
         '=SUMPRODUCT((Instructions!I6:I405="Allocated")*((Instructions!Q6:Q405="")+(Instructions!R6:R405="")>0))',
         '=IFERROR(MATCH(1,(Instructions!I6:I405="Allocated")*((Instructions!Q6:Q405="")+(Instructions!R6:R405="")>0),0)+5,"—")'),
        ("Council Response = Yes but no notes",
         '=SUMPRODUCT((Instructions!U6:U405="Yes")*(Instructions!V6:V405=""))',
         '=IFERROR(MATCH(1,(Instructions!U6:U405="Yes")*(Instructions!V6:V405=""),0)+5,"—")'),
        ("'Other' firm picked (needs A2Z follow-up)",
         '=COUNTIF(Instructions!J6:J405,"Other – advise A2Z Cloud")',
         '=IFERROR(MATCH("Other – advise A2Z Cloud",Instructions!J6:J405,0)+5,"—")'),
        ("'Other' client picked (needs A2Z follow-up)",
         '=COUNTIF(Instructions!K6:K405,"Other – advise A2Z Cloud")',
         '=IFERROR(MATCH("Other – advise A2Z Cloud",Instructions!K6:K405,0)+5,"—")'),
        ("'Other' investigator picked (needs A2Z follow-up)",
         '=COUNTIF(Instructions!Q6:Q405,"Other – advise A2Z Cloud")',
         '=IFERROR(MATCH("Other – advise A2Z Cloud",Instructions!Q6:Q405,0)+5,"—")'),
    ]
    for i, (issue, count_f, ex_f) in enumerate(issues):
        r = 11 + i
        ws[f"B{r}"] = issue
        ws[f"B{r}"].font = font_body
        ws[f"B{r}"].alignment = align_left
        ws[f"B{r}"].border = border_all
        ws[f"C{r}"] = count_f
        ws[f"C{r}"].font = font_body_bold
        ws[f"C{r}"].alignment = align_centre
        ws[f"C{r}"].border = border_all
        ws[f"D{r}"] = ex_f
        ws[f"D{r}"].font = font_body
        ws[f"D{r}"].alignment = align_centre
        ws[f"D{r}"].border = border_all
        ws.row_dimensions[r].height = 20

        # CF: highlight count cell in amber if > 0
        cf_rule = FormulaRule(
            formula=[f'$C{r}>0'], stopIfTrue=False, fill=fill_amber,
        )
        ws.conditional_formatting.add(f"C{r}:D{r}", cf_rule)

    # Footer note
    last_issue_row = 10 + len(issues)
    note_row = last_issue_row + 2
    ws.merge_cells(f"B{note_row}:D{note_row}")
    ws[f"B{note_row}"] = (
        "Tip: this dashboard updates live as you fill rows. The banner above turns green when there are no issues "
        "left and at least one row has been entered."
    )
    ws[f"B{note_row}"].font = font_descr
    ws[f"B{note_row}"].alignment = align_left
    ws.row_dimensions[note_row].height = 30

    ws.protection.sheet = True
    return ws


# ============================================================
#   ASSEMBLE
# ============================================================

def build():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Order matters — Cover first
    build_cover(wb)
    build_guide(wb)
    build_picklists(wb)  # build picklists early so defined names exist when referenced
    build_instructions(wb)
    build_portal_users(wb)
    build_readiness(wb)

    # Re-order so Cover is first, Picklists is last (and hidden)
    order = ["Cover", "Guide", "Instructions", "Portal Users", "Submission Readiness", "Picklists"]
    wb._sheets = [wb[name] for name in order]

    wb.active = 0  # start on Cover

    out_path_outputs = "/sessions/sweet-clever-archimedes/mnt/outputs/idi_migration/4593_ID_Inquiries_Migration_Intake_v5.xlsx"
    wb.save(out_path_outputs)
    print(f"Wrote {out_path_outputs}")

    # Also write to workspace folder
    out_path_ws = "/sessions/sweet-clever-archimedes/mnt/rg-system/4593_ID_Inquiries_Migration_Intake_v5.xlsx"
    wb.save(out_path_ws)
    print(f"Wrote {out_path_ws}")


if __name__ == "__main__":
    build()
